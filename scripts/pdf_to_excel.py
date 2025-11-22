#!/usr/bin/env python3
"""
PDF to Excel conversion with EXACT layout preservation.
Uses PDF → HTML → Excel approach:
1. Convert PDF to HTML with exact coordinates (using PyMuPDF)
2. Parse HTML with absolute positioning
3. Map text to Excel cells preserving exact layout
"""

import sys
import os
import re
import html as html_module
from collections import defaultdict
from pathlib import Path
import tempfile

try:
    import fitz  # PyMuPDF
except ImportError:
    print("ERROR: PyMuPDF library not installed. Install it with: pip install PyMuPDF", file=sys.stderr)
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl library not installed. Install it with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def parse_css_style(style_str):
    """Parse CSS style string and extract properties"""
    styles = {}
    for prop in style_str.split(';'):
        prop = prop.strip()
        if ':' in prop:
            key, value = prop.split(':', 1)
            styles[key.strip()] = value.strip()
    return styles


def extract_text_from_html(html_path: str):
    """
    Extract text elements with positions from HTML file.
    Returns list of text items with coordinates and formatting.
    """
    with open(html_path, 'r', encoding='utf-8') as f:
        html_content = f.read()
    
    text_items = []
    
    # Find all text elements with absolute positioning
    # Pattern: <div class="text-element" style="left: Xpx; top: Ypx; ...">text</div>
    pattern = r'<div class="text-element" style="([^"]+)">([^<]+)</div>'
    matches = re.findall(pattern, html_content)
    
    for style_str, text in matches:
        styles = parse_css_style(style_str)
        
        # Extract position
        left = float(styles.get('left', '0').replace('px', ''))
        top = float(styles.get('top', '0').replace('px', ''))
        
        # Extract font properties
        font_size = float(styles.get('font-size', '12').replace('px', ''))
        font_family = styles.get('font-family', 'Arial')
        color = styles.get('color', 'rgb(0, 0, 0)')
        is_bold = 'bold' in styles.get('font-weight', '').lower()
        is_italic = 'italic' in styles.get('font-style', '').lower()
        
        # Unescape HTML entities
        text = html_module.unescape(text)
        
        if text.strip():
            text_items.append({
                'text': text,
                'x': left,
                'y': top,
                'font_size': font_size,
                'font_family': font_family,
                'color': color,
                'bold': is_bold,
                'italic': is_italic
            })
    
    return text_items


def pdf_to_html_temp(pdf_path: str) -> str:
    """
    Convert PDF to HTML temporarily and return HTML path.
    Uses the same approach as pdf_to_html.py
    """
    # Create temporary HTML file
    temp_dir = tempfile.gettempdir()
    unique_id = f"pdf_html_{os.getpid()}_{id(pdf_path)}"
    html_path = os.path.join(temp_dir, f"{unique_id}.html")
    
    # Open PDF
    doc = fitz.open(pdf_path)
    
    # Build HTML
    html_content = []
    html_content.append('<!DOCTYPE html>')
    html_content.append('<html lang="en">')
    html_content.append('<head>')
    html_content.append('    <meta charset="UTF-8">')
    html_content.append('    <style>')
    html_content.append('        .text-element {')
    html_content.append('            position: absolute;')
    html_content.append('            white-space: pre;')
    html_content.append('            line-height: 1.0;')
    html_content.append('        }')
    html_content.append('    </style>')
    html_content.append('</head>')
    html_content.append('<body>')
    
    # Process each page
    for page_num in range(len(doc)):
        page = doc[page_num]
        page_rect = page.rect
        page_width = page_rect.width
        page_height = page_rect.height
        
        html_content.append(f'    <div class="page" style="width: {page_width}px; height: {page_height}px;">')
        
        # Extract text blocks with positions
        text_dict = page.get_text("dict")
        
        for block in text_dict.get("blocks", []):
            if "lines" not in block:
                continue
            
            for line in block.get("lines", []):
                for span in line.get("spans", []):
                    text = span.get("text", "").strip()
                    if not text:
                        continue
                    
                    bbox = span.get("bbox", [0, 0, 0, 0])
                    x0, y0, x1, y1 = bbox
                    
                    font_size = span.get("size", 12)
                    font_name = span.get("font", "Arial")
                    flags = span.get("flags", 0)
                    
                    is_bold = flags & 16
                    is_italic = flags & 1
                    
                    color = span.get("color", 0)
                    r = (color >> 16) & 0xFF
                    g = (color >> 8) & 0xFF
                    b_val = color & 0xFF
                    color_str = f"rgb({r}, {g}, {b_val})"
                    
                    text_escaped = html_module.escape(text)
                    
                    style_parts = [
                        f"left: {x0}px",
                        f"top: {y0}px",
                        f"font-size: {font_size}px",
                        f"font-family: '{font_name}', Arial, sans-serif",
                        f"color: {color_str}"
                    ]
                    
                    if is_bold:
                        style_parts.append("font-weight: bold")
                    if is_italic:
                        style_parts.append("font-style: italic")
                    
                    style_str = "; ".join(style_parts)
                    html_content.append(f'        <div class="text-element" style="{style_str}">{text_escaped}</div>')
        
        html_content.append('    </div>')
    
    html_content.append('</body>')
    html_content.append('</html>')
    
    # Write HTML file
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(html_content))
    
    doc.close()
    return html_path


def group_text_by_lines(text_items, line_tolerance=5):
    """Group text items into lines based on Y position"""
    if not text_items:
        return []
    
    # Sort by Y position
    sorted_items = sorted(text_items, key=lambda item: (item['y'], item['x']))
    
    lines = []
    current_line = []
    current_y = None
    
    for item in sorted_items:
        y_pos = item['y']
        
        if current_y is None or abs(y_pos - current_y) <= line_tolerance:
            current_line.append(item)
            if current_y is None:
                current_y = y_pos
            else:
                current_y = (current_y + y_pos) / 2
        else:
            if current_line:
                current_line.sort(key=lambda item: item['x'])
                lines.append({
                    'items': current_line,
                    'y': current_y
                })
            current_line = [item]
            current_y = y_pos
    
    if current_line:
        current_line.sort(key=lambda item: item['x'])
        lines.append({
            'items': current_line,
            'y': current_y
        })
    
    return lines


def detect_columns(lines, page_width):
    """Detect column boundaries based on text X positions"""
    if not lines:
        return []
    
    x_positions = []
    for line in lines:
        for item in line['items']:
            x_positions.append(item['x'])
    
    if not x_positions:
        return []
    
    x_positions = sorted(set(x_positions))
    
    # Group nearby X positions into columns
    columns = []
    current_col = None
    threshold = page_width * 0.05  # 5% of page width
    
    for x in x_positions:
        if current_col is None:
            current_col = x
        elif abs(x - current_col) <= threshold:
            current_col = (current_col + x) / 2
        else:
            columns.append(current_col)
            current_col = x
    
    if current_col is not None:
        columns.append(current_col)
    
    return sorted(set(columns))


def get_column_index(x, columns):
    """Get the column index for a given X position"""
    if not columns:
        return 1
    
    for i, col_x in enumerate(columns):
        if x < col_x:
            return i + 1
    
    return len(columns) + 1


def convert_pdf_to_excel(pdf_path: str, excel_path: str) -> bool:
    """
    Convert PDF to Excel with exact layout preservation.
    Uses PDF → HTML → Excel approach.
    """
    temp_html_path = None
    
    try:
        if not os.path.exists(pdf_path):
            print(f"ERROR: PDF file not found: {pdf_path}", file=sys.stderr)
            return False
        
        print(f"Step 1: Converting PDF to HTML with exact layout...", file=sys.stderr)
        temp_html_path = pdf_to_html_temp(pdf_path)
        
        if not os.path.exists(temp_html_path):
            print(f"ERROR: Failed to create HTML file", file=sys.stderr)
            return False
        
        print(f"Step 2: Extracting text with positions from HTML...", file=sys.stderr)
        text_items = extract_text_from_html(temp_html_path)
        
        if not text_items:
            print("Warning: No text found in PDF.", file=sys.stderr)
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=1).value = "No extractable text found in PDF"
            wb.save(excel_path)
            return True
        
        print(f"  Extracted {len(text_items)} text items", file=sys.stderr)
        
        # Group text into lines
        lines = group_text_by_lines(text_items, line_tolerance=8)
        print(f"  Grouped into {len(lines)} lines", file=sys.stderr)
        
        # Create Excel workbook
        wb = Workbook()
        wb.remove(wb.active)
        
        # Process each page (for now, treat all as one page)
        # In future, we can detect page breaks from HTML
        ws = wb.create_sheet(title="Page_1")
        
        # Detect columns
        if lines:
            # Estimate page width from max X position
            max_x = max(item['x'] for line in lines for item in line['items'])
            page_width = max_x + 100  # Add margin
            columns = detect_columns(lines, page_width)
            print(f"  Detected {len(columns)} column(s)", file=sys.stderr)
        else:
            columns = []
        
        # Set reasonable column widths
        max_cols = min(len(columns) + 1, 20)
        for col in range(1, max_cols + 1):
            ws.column_dimensions[get_column_letter(col)].width = 30
        
        current_row = 1
        
        # Write text to Excel
        for line in lines:
            # Group items by column
            items_by_col = defaultdict(list)
            
            for item in line['items']:
                col_idx = get_column_index(item['x'], columns)
                items_by_col[col_idx].append(item)
            
            # Write items to their respective columns
            for col_idx, items in sorted(items_by_col.items()):
                # Combine text in this column
                combined_text = ' '.join([item['text'] for item in items])
                
                # Get formatting from first item (or combine)
                first_item = items[0]
                font_size = first_item['font_size']
                is_bold = any(item['bold'] for item in items)
                is_italic = any(item['italic'] for item in items)
                
                # Write to Excel
                cell = ws.cell(row=current_row, column=col_idx)
                if cell.value:
                    cell.value = str(cell.value) + ' ' + combined_text
                else:
                    cell.value = combined_text
                
                # Apply formatting
                cell.font = Font(
                    size=int(font_size),
                    bold=is_bold,
                    italic=is_italic
                )
                cell.alignment = Alignment(
                    vertical='top',
                    horizontal='left',
                    wrap_text=True
                )
            
            # Set row height
            ws.row_dimensions[current_row].height = 20
            current_row += 1
        
        # Save workbook
        wb.save(excel_path)
        
        # Clean up temp HTML file
        if temp_html_path and os.path.exists(temp_html_path):
            try:
                os.unlink(temp_html_path)
            except:
                pass
        
        if os.path.exists(excel_path):
            print(f"SUCCESS: Converted {pdf_path} to {excel_path}", file=sys.stderr)
            return True
        else:
            print(f"ERROR: Output file was not created: {excel_path}", file=sys.stderr)
            return False
            
    except Exception as e:
        print(f"ERROR: Conversion failed: {str(e)}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        
        # Clean up temp HTML file
        if temp_html_path and os.path.exists(temp_html_path):
            try:
                os.unlink(temp_html_path)
            except:
                pass
        
        return False


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python pdf_to_excel.py <input_pdf> <output_xlsx>", file=sys.stderr)
        sys.exit(1)
    
    input_pdf = sys.argv[1]
    output_excel = sys.argv[2]
    
    success = convert_pdf_to_excel(input_pdf, output_excel)
    sys.exit(0 if success else 1)
